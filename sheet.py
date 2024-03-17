import xlwt
import openpyxl
from xlutils.copy import copy
import os
import datetime
import xlsxwriter

st_name = 'Aashish'
def mark_present(st_name):
	directory = 'C:/Users/ameya/Deep learning/attendance_deep_learning/scripts_used/sample/output/'
	names = os.listdir(directory)
	print(names)

	sub = 'SAMPLE'
	
	if not os.path.exists('C:/Users/ameya/Deep learning/attendance_deep_learning/scripts_used/sample/attendance/' + sub + '.xlsx'):
		count = 2
		workbook = xlsxwriter.Workbook('C:/Users/ameya/Deep learning/attendance_deep_learning/scripts_used/sample/attendance/' + sub + '.xlsx')
		print("Creating Spreadsheet with Title: " + sub)
		sheet = workbook.add_worksheet() 
		for i in names:
			sheet.write(count, 0, i)
			count += 1
		workbook.close() 

	rb = openpyxl.load_workbook('C:/Users/ameya/Deep learning/attendance_deep_learning/scripts_used/sample/attendance/' + sub + '.xlsx')
	wb = rb
	sheet = wb.active
	sheet.cell(1,1,str(datetime.datetime.now()))


	count = 2
	for i in names:
		if i in st_name:
			sheet.cell(count, 1, 'P')
		else:
			sheet.cell(count, 1, 'A')
		sheet.cell(count, 1, i)
		count += 1

	wb.save('C:/Users/ameya/Deep learning/attendance_deep_learning/scripts_used/sample/attendance/' + sub + '.xlsx')


mark_present(st_name)
